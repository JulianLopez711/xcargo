import os
import smtplib
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import locale
import logging

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('reporte_pagos')

# Establecer el locale para nombres de meses en español
try:
    locale.setlocale(locale.LC_TIME, "es_CO.UTF-8")
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")  # Fallback para España
    except locale.Error:
        logger.warning("No se pudo establecer locale español, usando el predeterminado")

# Cargar variables de entorno
load_dotenv()

EMAIL = os.getenv("EMAIL_SENDER")
PASSWORD = os.getenv("EMAIL_PASSWORD")

# Emails de clientes
DESTINOS = {
    "Dropi": os.getenv("EMAIL_CLIENTE_DROPI"),
    "Dafiti": os.getenv("EMAIL_CLIENTE_DAFITI"),
    "Trady": os.getenv("EMAIL_CLIENTE_TRADY"),
}

def generar_excel_entregas(entregas: list, cliente: str) -> bytes:
    """
    Genera un archivo Excel con el detalle de las entregas realizadas.
    
    Args:
        entregas (list): Lista de diccionarios con información de entregas
        cliente (str): Nombre del cliente para personalizar el Excel
        
    Returns:
        bytes: Contenido del archivo Excel en formato bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Entregas"
    
    # Estilos
    titulo_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Título del reporte
    fecha_actual = datetime.now().strftime("%d de %B de %Y")
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Reporte de entregas - {cliente} - {fecha_actual}"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Ajustar ancho de columnas
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18
    
    # Encabezados con estilo
    headers = ["Tracking", "Fecha", "Tipo", "Cliente", "Valor"]
    ws.append(headers)
    
    # Aplicar estilos a los encabezados
    for col in range(1, 6):
        cell = ws.cell(row=2, column=col)
        cell.font = titulo_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Agregar datos
    row_num = 3
    total_valor = 0
    for entrega in entregas:
        ws.append([
            entrega["tracking"],
            entrega["fecha"],
            entrega["tipo"],
            entrega["cliente"],
            entrega["valor"]
        ])
        
        # Dar formato a la celda de valor como moneda
        ws[f'E{row_num}'].number_format = '"$"#,##0.00'
        
        # Aplicar bordes a todas las celdas
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = border
            
        total_valor += float(entrega["valor"])
        row_num += 1
    
    # Agregar fila de totales
    ws.merge_cells(f'A{row_num}:D{row_num}')
    ws[f'A{row_num}'] = "TOTAL"
    ws[f'A{row_num}'].font = Font(bold=True)
    ws[f'A{row_num}'].alignment = Alignment(horizontal='right')
    
    ws[f'E{row_num}'] = total_valor
    ws[f'E{row_num}'].font = Font(bold=True)
    ws[f'E{row_num}'].number_format = '"$"#,##0.00'
    
    for col in range(1, 6):
        ws.cell(row=row_num, column=col).border = border
        ws.cell(row=row_num, column=col).fill = PatternFill(start_color="E9FDF4", end_color="E9FDF4", fill_type="solid")
    
    # Guardar el archivo
    archivo_virtual = BytesIO()
    wb.save(archivo_virtual)
    archivo_virtual.seek(0)
    return archivo_virtual.read()

def obtener_periodo_reporte():
    """
    Obtiene el periodo de reporte: desde una semana antes hasta hoy.
    
    Returns:
        tuple: (fecha_inicio, fecha_fin) en formato legible
    """
    fecha_fin = datetime.now()
    fecha_inicio = fecha_fin - timedelta(days=7)  # Periodo de una semana
    
    return (
        fecha_inicio.strftime("%d de %B"),
        fecha_fin.strftime("%d de %B de %Y")
    )

def enviar_correo_pago(cliente: str, total: float, entregas: list, nombre_archivo: str, archivo_bytes: bytes):
    """
    Envía un correo con el reporte de pagos y los archivos adjuntos.
    
    Args:
        cliente (str): Nombre del cliente destinatario
        total (float): Monto total de las entregas
        entregas (list): Lista de diccionarios con información de entregas
        nombre_archivo (str): Nombre del archivo comprobante a adjuntar
        archivo_bytes (bytes): Contenido del archivo comprobante en formato bytes
        
    Returns:
        bool: True si el envío fue exitoso, False en caso contrario
    """
    try:
        # Obtener el periodo del reporte
        fecha_inicio, fecha_fin = obtener_periodo_reporte()
        
        # Crear el mensaje
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Reporte de recaudo {cliente} - {fecha_fin}"
        msg["From"] = f"XCARGO <{EMAIL}>"
        
        # Determinar el destinatario
        destinatario = DESTINOS.get(cliente)
        if not destinatario:
            logger.warning(f"No se encontró email para el cliente {cliente}, usando remitente como destinatario")
            destinatario = EMAIL
            
        msg["To"] = destinatario
        
        # Crear la versión en texto plano
        texto_plano = f"""
Buenos días, estimados {cliente}:

Adjunto el reporte correspondiente al recaudo por concepto de Cash on Delivery del periodo comprendido entre el {fecha_inicio} y el {fecha_fin}.

RESUMEN:
- Total recaudado: ${total:,.2f}
- Cantidad de entregas: {len(entregas)}

Se incluyen en este envío los soportes correspondientes y la relación en Excel con el detalle de lo recaudado.

Saludos cordiales,
Equipo XCARGO
        """
        
        # Crear la versión en HTML
        texto_html = f"""
<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Reporte de Recaudo</title>
  <style>
    body, html {{
      font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
      line-height: 1.6;
      color: #333;
      margin: 0;
      padding: 0;
    }}
    .container {{
      max-width: 600px;
      margin: 0 auto;
      padding: 20px;
    }}
    .header {{
      border-bottom: 3px solid #10B981;
      padding-bottom: 15px;
      margin-bottom: 20px;
    }}
    .logo {{
      font-size: 24px;
      font-weight: bold;
      color: #10B981;
    }}
    .content {{
      margin-bottom: 25px;
    }}
    .summary {{
      background-color: #f0fdf4;
      border-left: 4px solid #10B981;
      padding: 15px;
      margin: 20px 0;
    }}
    .summary-title {{
      font-weight: bold;
      margin-bottom: 10px;
    }}
    .summary-item {{
      margin: 8px 0;
    }}
    .footer {{
      margin-top: 30px;
      padding-top: 15px;
      border-top: 1px solid #e5e7eb;
      font-size: 12px;
      color: #6b7280;
      text-align: center;
    }}
    .highlight {{
      color: #10B981;
      font-weight: bold;
    }}
  </style>
</head>
<body>
  <div class="container">
    <div class="header">
      <div class="logo">XCARGO</div>
    </div>
    
    <div class="content">
      <p>Buenos días, estimados <strong>{cliente}</strong>:</p>
      
      <p>Adjunto el reporte correspondiente al recaudo por concepto de <span class="highlight">Cash on Delivery</span> 
      del periodo comprendido entre el <strong>{fecha_inicio}</strong> y el <strong>{fecha_fin}</strong>.</p>
      
      <div class="summary">
        <div class="summary-title">RESUMEN DEL REPORTE:</div>
        <div class="summary-item">• Total recaudado: <span class="highlight">${total:,.2f}</span></div>
        <div class="summary-item">• Cantidad de entregas: <span class="highlight">{len(entregas)}</span></div>
      </div>
      
      <p>Se incluyen en este envío los soportes correspondientes y la relación en Excel con el detalle de lo recaudado.</p>
      
      <p>Por favor, no dude en contactarnos si necesita información adicional.</p>
      
      <p>Saludos cordiales,<br>
      <strong>Equipo XCARGO</strong></p>
    </div>
    
    <div class="footer">
      © 2025 XCARGO. Todos los derechos reservados.<br>
      Este correo y sus adjuntos contienen información confidencial.
    </div>
  </div>
</body>
</html>
        """
        
        # Adjuntar versiones de texto y HTML
        parte_texto = MIMEText(texto_plano, "plain")
        parte_html = MIMEText(texto_html, "html")
        
        msg.attach(parte_texto)
        msg.attach(parte_html)
        
        # Adjuntar comprobante
        adjunto1 = MIMEBase("application", "octet-stream")
        adjunto1.set_payload(archivo_bytes)
        encoders.encode_base64(adjunto1)
        adjunto1.add_header(
            "Content-Disposition", 
            f"attachment; filename={nombre_archivo}"
        )
        msg.attach(adjunto1)
        
        # Generar y adjuntar Excel
        excel_bytes = generar_excel_entregas(entregas, cliente)
        adjunto2 = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        adjunto2.set_payload(excel_bytes)
        encoders.encode_base64(adjunto2)
        adjunto2.add_header(
            "Content-Disposition", 
            f"attachment; filename=Entregas_{cliente}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        msg.attach(adjunto2)
        
        # Enviar correo
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL, PASSWORD)
            smtp.send_message(msg)
            
        logger.info(f"Correo enviado exitosamente a {cliente} ({destinatario})")
        return True
            
    except Exception as e:
        logger.error(f"Error al enviar correo a {cliente}: {str(e)}")
        return False

